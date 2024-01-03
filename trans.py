#!/usr/bin/env python3
from openpyxl import load_workbook
from pypinyin import lazy_pinyin
import json
import sys
import os
import hashlib

def parseInt(s):
    try:
        if s is not None:
            if s.strip():
                return int(s.strip())
        return 0
    except:
        print("failed:",s)
        return 0
def parse_xlsx(xlsx_file: str):
    d = []
    wb = load_workbook(xlsx_file)
    ws = wb.active
    i = 0
    totalTags = set()
    for row in ws[f'2:{ws.max_row}']:
        if row[1].value is None:
            break
        money = parseInt(row[0].value)
        song = str(row[1].value).strip()
        singer = row[2].value and str(row[2].value).strip() or ''
        link = row[3].value and (row[3].value).strip() or ''
        tags = [lazy_pinyin(song[:1])[0].upper()[:1]]
        tags += ['付费'] if money else []
        tags += ['视频'] if link else []
        tags += row[4].value and row[4].value.strip() and [x.upper() for x in str(row[4].value).replace('，', ',').strip().split(',')] or []
        if "学习中" in tags:
            continue
        remark = row[5].value and str(row[5].value).strip() or ''
        remark += f' {"".join(lazy_pinyin(song))} {"".join(lazy_pinyin(singer))}'
        remark = remark.strip()
        d.append({
            'key': i,
            'money': money,
            'song': song,
            'singer': singer,
            'link': link,
            'tags': tags,
            'remark': remark,
        })
        for tag in tags:
            totalTags.add(tag)
        i += 1
    return d, totalTags


if __name__ == '__main__':
    data, allTagsSet = parse_xlsx(sys.argv[1] if len(sys.argv) > 1 else '小茄歌单.xlsx')
    allTagsList = list(allTagsSet)
    allTagsList.sort()
    jdata = json.dumps({'data': data, 'tags': allTagsList},ensure_ascii=False, separators=(',', ':'))
    with open(os.path.join('src', 'assets', 'data.json'), 'w', encoding='utf-8') as f:
        f.write(jdata)
#     print(json.dumps({'data': data}, indent=4))
    hasher = hashlib.md5()
    hasher.update(jdata.encode('utf-8'))
    with open(os.path.join('public', 'data_hash.txt'), 'w', encoding='utf-8') as f:
        f.write(hasher.hexdigest())
    print(f"md5:{hasher.hexdigest()}")
    print('OK')
